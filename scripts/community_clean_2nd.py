#!/usr/bin/env python3
"""
对 community_cleaned.xlsx 做二次清洗,把人工复核后仍残留的问题再过一遍。

输入输出:
  - 输入: community_cleaned.xlsx(由 community-clean.py 生成,人工已手动改正过)
  - 输出: 同名文件,覆写

清洗规则(顺序执行):
  1. 「地址」 列转 JSON 数组
     - 已是 `[...]` 形式:解析为 list
     - 否则按 [,,;,;、,] 拆分,空段丢弃
     - 解析后转回 json.dumps 输出
  2. 共享前缀地址修复
     形如 [联青路50弄, 51弄, 135弄, 180弄, 198弄, 花王路450弄] —— 中间几条只是
     「数字+弄/号」,缺前缀,应该共享上一条路名。
     规则:如果当前元素不包含路名要素(路/街/道/巷/村),且上一条以路名要素结尾,
     就把上一条的路名前缀补上;遇到含路名的元素,重置前缀。
  3. 「名称」 长度 >= 6 且末位是「小区」 → 剥掉末位「小区」
     例:「上海阳城小区」 → 「上海阳城」
     例外:「S32小区」(len=5)、「永德小区1」(「小区」 不在末尾)不动。
  4. 「名称」 里纯阿拉伯数字 + 「期/村/街坊」 单位 → 数字转中文
     例:「万源新城1-3期」 → 「万源新城一~三期」
     例外:
       - 数字前是字母(S32 / A1 / E1)整段不动
       - 括号内的数字不动(「(15-01地块)」 保留)
       - 仅识别 期/村/街坊/号/弄/支弄/组 这几个单位

依赖: pip3 install openpyxl
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC_XLSX = ROOT / "community_cleaned.xlsx"

# 1a. 裸字符串场景下的地址分隔符:逗号/分号/顿号
_ADDR_DELIM_RE = re.compile(r"[,，;；、]")
# 1b. 已经是 `[...]` list 形式时的分隔符:只切逗号/分号,**不切顿号**。
#     理由:list 内部顿号(如 「吴中路907、909号」)通常是同段路名内部的多个门牌,
#     不该被拆成多条独立地址 —— 用户原例就是这种场景。
_ADDR_DELIM_LIST_RE = re.compile(r"[,，;；]")
# 2. 共享前缀修复:用于识别「路名要素」(决定前缀是否需要补)
_ROAD_KEYWORDS = ("路", "街", "道", "巷", "村", "组")
#   「数字 + 单位」 元素:开头是纯数字,后面紧跟路名/序号单位。
#   例:「51弄」 / 「135弄」 / 「198弄」 / 「38支弄」 / 「250号」
_NUM_UNIT_RE = re.compile(
    r"^\d+(?:支弄|弄|号|期|村|街坊|组)(?:.*)$"
)
# 4. 数字 → 中文
_DIGIT_TRANS_UNITS = ("期", "村", "街坊")
#   单数字 0-9 的中文表
_DIGIT_CN = "零一二三四五六七八九"
#   范围分隔符:ASCII -, 全角 —, 中文 ～
_RANGE_SEP_RE = re.compile(r"[-—–~～]")


# ————————————————————— 1. 地址 → JSON —————————————————————

def normalize_addr_cell(v) -> str | None:
    """把「地址」 单元格值规范成 JSON 数组字符串。

    - None / 空 → None
    - 已是 list → json.dumps
    - 已是 str:
        - 能 json.loads 成 list → 重新 dumps(确保格式统一)
        - 否则按 [,,;;,,] 拆成 list,空段丢弃
    """
    if v is None:
        return None
    if isinstance(v, (list, tuple)):
        items = [str(x).strip() for x in v if str(x).strip()]
        return json.dumps(items, ensure_ascii=False) if items else None
    if isinstance(v, (int, float)):
        v = str(v)
    s = str(v).strip()
    if not s:
        return None
    # 尝试直接 parse
    if s.startswith("[") and s.endswith("]"):
        try:
            arr = json.loads(s)
            if isinstance(arr, list):
                return json.dumps(
                    [str(x).strip() for x in arr if str(x).strip()],
                    ensure_ascii=False,
                ) or None
        except json.JSONDecodeError:
            # 不是合法 JSON(比如 "[联青路50弄, 51弄]" 没引号),剥方括号后按分隔符切
            # list 形式:**不切顿号**,避免 「吴中路907、909号」 被拆
            inner = s[1:-1].strip()
            if inner:
                parts = [p.strip() for p in _ADDR_DELIM_LIST_RE.split(inner) if p.strip()]
                return json.dumps(parts, ensure_ascii=False) if parts else None
            return None
    # 裸字符串场景:按所有分隔符(逗号/分号/顿号)切
    parts = [p.strip() for p in _ADDR_DELIM_RE.split(s) if p.strip()]
    return json.dumps(parts, ensure_ascii=False) if parts else None


# ————————————————————— 2. 共享前缀地址修复 —————————————————————

def _has_road_keyword(s: str) -> bool:
    return any(k in s for k in _ROAD_KEYWORDS)


def _starts_with_digit_unit(s: str) -> bool:
    """「数字 + 单位」 开头,且不含路名要素。"""
    if _has_road_keyword(s):
        return False
    return bool(re.match(r"^\d", s)) and bool(_NUM_UNIT_RE.match(s))


def _extract_road_prefix(prev: str) -> str | None:
    """从上一条地址里抽出「路名」 前缀,供下一条补全。

    抽取策略:从左到右扫描,直到遇到第一个数字/单位(弄/号/支弄/期/村/组)为止,
    返回该前缀(含路名)。
    例:「联青路50弄」 -> 「联青路」
       「花王路450弄」 -> 「花王路」
       「吴中路907、909号」 -> 「吴中路」
    """
    for i, ch in enumerate(prev):
        if ch.isdigit() or ch in "弄号期村组支":
            return prev[:i] if i > 0 else None
    return None


def repair_shared_prefix_addresses(addresses: list[str]) -> list[str]:
    """共享前缀地址修复。

    见模块 docstring 规则 2。
    """
    if not addresses:
        return addresses
    out: list[str] = []
    prefix: str | None = None
    for cur in addresses:
        cur = cur.strip()
        if not cur:
            continue
        if _starts_with_digit_unit(cur):
            # 缺路名,需要补前缀
            if prefix is None:
                # 没有可继承的前缀,保留原样
                out.append(cur)
                # 不更新 prefix
                continue
            out.append(prefix + cur)
        else:
            out.append(cur)
            prefix = _extract_road_prefix(cur)
    return out


# ————————————————————— 3. 剥「小区」 —————————————————————

def strip_xiaoqu_suffix(name: str) -> str:
    """名称长度 >= 6 且末尾 2 字是「小区」 → 剥掉。

    例:「上海阳城小区」 (len=6) → 「上海阳城」
       「S32小区」 (len=5) → 不动
       「永德小区1」 (不以「小区」 结尾) → 不动
    """
    if not name:
        return name
    s = str(name)
    if len(s) >= 6 and s.endswith("小区"):
        return s[:-2]
    return s


# ————————————————————— 4. 数字 → 中文 —————————————————————

def _digit_to_cn(d: str) -> str:
    """纯阿拉伯数字字符串 -> 中文数字字符串。

    例:「1」 -> 「一」,「23」 -> 「二三」(逐字翻译,符合「X期/村/街坊」 场景)。
    """
    return "".join(_DIGIT_CN[int(ch)] if ch.isdigit() else ch for ch in d)


def _maybe_translate_digits_in_unit(name: str) -> str:
    r"""把「名称」 里 [期/村/街坊] 单位前的纯数字翻译成中文。

    规则:
      - 仅匹配 `\d+(?:\s*[-—–~～]\s*\d+)?[期村街坊]` 这种完整单位短语
      - 数字前面必须是「非字母字符 或 字符串开头」,字母紧贴数字的(如 S32, E1)整段不动
      - 范围 1-3 / 6—13 转 `一~三 / 六~一三`
      - 括号内的不处理
    """
    if not name:
        return name

    out: list[str] = []
    i = 0
    n = len(name)
    while i < n:
        # 括号内透传
        if name[i] in "（(":
            close = "）)"["（(".index(name[i])]
            j = name.find(close, i + 1)
            if j == -1:
                # 括号不闭合,当普通字符处理
                out.append(name[i])
                i += 1
                continue
            out.append(name[i : j + 1])
            i = j + 1
            continue

        # 匹配 \d+(range)?[期村街坊]
        m = re.match(r"(\d+)((?:\s*[-—–~～]\s*\d+)?)([期村街坊])", name[i:])
        if not m:
            out.append(name[i])
            i += 1
            continue

        digit_part, range_part, unit = m.group(1), m.group(2), m.group(3)
        start_in_name = i
        end_in_name = i + len(m.group(0))

        # 数字前面必须是「字符串开头 / 非 ASCII 字母」,ASCII 字母紧贴数字的
        # (如 S32 / E1)整段不动;中文不算字母,正常翻译。
        if start_in_name > 0:
            prev_ch = name[start_in_name - 1]
            if prev_ch.isascii() and prev_ch.isalpha():
                # ASCII 字母 + 数字(如 S32 / E1),跳过翻译
                out.append(name[i])
                i += 1
                continue

        # 翻译数字
        cn_digit = _digit_to_cn(digit_part)
        if range_part:
            # 范围分隔符统一为 ~
            nums = re.split(r"\s*[-—–~～]\s*", range_part.strip())
            cn_nums = [_digit_to_cn(x) for x in nums if x]
            translated = f"{cn_digit}~{'~'.join(cn_nums)}{unit}"
        else:
            translated = f"{cn_digit}{unit}"
        out.append(translated)
        i = end_in_name
    return "".join(out)


# ————————————————————— 主流程 —————————————————————

def process_workbook(path: Path = SRC_XLSX) -> None:
    """读 xlsx,逐行清洗,覆写回 xlsx。

    约定列:
      - 第 1 列:名称
      - 第 8 列:地址(由人工核对后保留的 8 列结构)
    列数若不符,脚本尝试按表头查找列名「名称」 / 「地址」,找不到则报错。
    """
    if not path.exists():
        raise FileNotFoundError(f"找不到输入文件: {path}")

    wb = openpyxl.load_workbook(path)
    if "communities" not in wb.sheetnames:
        raise RuntimeError(f"{path} 缺 communities sheet")
    ws = wb["communities"]

    header = [c.value for c in ws[1]]
    try:
        name_col = header.index("名称") + 1
        addr_col = header.index("地址") + 1
    except ValueError as e:
        raise RuntimeError(f"表头找不到「名称」/「地址」 列: {e}") from e

    name_changes = 0
    addr_changes = 0
    addr_repairs = 0
    total = 0

    for r in range(2, ws.max_row + 1):
        total += 1
        # 名称清洗:剥小区 + 数字转中文
        old_name = ws.cell(r, name_col).value
        if old_name:
            step1 = strip_xiaoqu_suffix(str(old_name))
            step2 = _maybe_translate_digits_in_unit(step1)
            if step2 != old_name:
                ws.cell(r, name_col).value = step2
                name_changes += 1

        # 地址清洗:转 JSON + 共享前缀修复
        old_addr_raw = ws.cell(r, addr_col).value
        new_addr_json = normalize_addr_cell(old_addr_raw)
        if new_addr_json:
            # 解析为 list 做共享前缀修复,再 dumps
            try:
                addrs = json.loads(new_addr_json)
            except json.JSONDecodeError:
                addrs = []
            repaired = repair_shared_prefix_addresses(addrs)
            new_json2 = json.dumps(repaired, ensure_ascii=False)
            if new_json2 != new_addr_json:
                addr_repairs += 1
            final = new_json2
        else:
            final = None
        if final != (None if old_addr_raw is None else str(old_addr_raw)):
            ws.cell(r, addr_col).value = final
            addr_changes += 1

    wb.save(path)
    print(f"清洗完成 -> {path}")
    print(f"  总行数: {total}")
    print(f"  名称变化: {name_changes} (剥小区 + 数字转中文)")
    print(f"  地址列变化: {addr_changes}")
    print(f"  共享前缀地址修复: {addr_repairs}")


if __name__ == "__main__":
    process_workbook()
