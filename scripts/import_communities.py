#!/usr/bin/env python3
"""
把 community_cleaned.xlsx 导入到 communities 表。

策略:
  1. 读 xlsx,每行映射到 (name, region_id, alias, address, subarea, status)
  2. xlsx 的「区划ID」 是 regions.code(12 位 addressStandardCode),通过 regions.code
     查 regions.id(cuid) 后写入;匹配不到的 region_id 留 NULL。
  3. 清空 communities 表后再写(用户确认覆盖模式)。
  4. status 默认 1(启用)。
  5. subarea:DB 当前没有该列(已 DROP),脚本读 xlsx 这一列但不写入,
     等后续 DDL 加回后,只需在本脚本里把 subarea 加到 INSERT 字段即可。
  6. geom:DB 是 GEOMCOLLECTION(空间类型),暂不写,统一 NULL。

用法:
    python3 scripts/import_communities.py
依赖:
    pip3 install openpyxl pymysql
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
import pymysql

ROOT = Path(__file__).resolve().parent.parent
SRC_XLSX = ROOT / "community_cleaned.xlsx"
ENV_FILE = ROOT / ".env"


def load_database_url() -> str:
    text = ENV_FILE.read_text(encoding="utf-8")
    m = re.search(r'DATABASE_URL="([^"]+)"', text)
    if not m:
        raise RuntimeError(f"{ENV_FILE} 里没找到 DATABASE_URL")
    return m.group(1)


def connect():
    url = urlparse(load_database_url())
    return pymysql.connect(
        host=url.hostname,
        port=url.port or 3306,
        user=url.username,
        password=url.password or "",
        database=url.path.lstrip("/"),
        charset="utf8mb4",
        autocommit=False,
    )


def parse_json_cell(v) -> object | None:
    """xlsx 中 JSON 字符串单元格 → Python 对象(无法解析返回 None)。

    兼容:
      - None / 空字符串 → None
      - 已经是 list → 原样返回
      - JSON 字符串 → json.loads
    """
    if v is None:
        return None
    if isinstance(v, (list, tuple)):
        return list(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        return None


def import_communities() -> None:
    if not SRC_XLSX.exists():
        raise FileNotFoundError(f"找不到输入文件: {SRC_XLSX}")

    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    if "communities" not in wb.sheetnames:
        raise RuntimeError(f"{SRC_XLSX} 缺 communities sheet")
    ws = wb["communities"]

    header = [c.value for c in ws[1]]
    # 列定位(从表头找,而不是固定索引,兼容顺序调整)
    def find_col(name: str) -> int | None:
        try:
            return header.index(name) + 1
        except ValueError:
            return None

    col_name = find_col("名称")
    col_alias = find_col("别名")
    col_region_id = find_col("区划ID")  # 12 位 addressStandardCode
    col_region_name = find_col("匹配区划名称")  # 调试日志用
    col_address = find_col("地址")
    col_subarea = find_col("子区域")
    col_status = find_col("匹配置信度")  # 误用,实际 status 默认 1;先这么写,后续按需调整

    if col_name is None or col_address is None:
        raise RuntimeError("表头缺「名称」/「地址」列")

    print(
        f"列定位: name={col_name} alias={col_alias} regionId={col_region_id} "
        f"regionName={col_region_name} address={col_address} subarea={col_subarea} status={col_status}"
    )

    conn = connect()
    try:
        with conn.cursor() as cur:
            # 0) 先把 regions 表的 code → id 映射拉出来
            cur.execute("SELECT id, code FROM regions")
            code_to_id: dict[str, str] = {row[1]: row[0] for row in cur.fetchall()}
            print(f"regions 表加载: {len(code_to_id)} 条 code → id 映射")

            # 1) 清空 communities 表
            cur.execute("DELETE FROM communities")
            deleted = cur.rowcount
            print(f"清空 communities 表: 删除 {deleted} 条旧记录")

            # 2) 逐行读取 + INSERT
            inserted = 0
            skipped = 0
            region_unmatched = 0
            for r in range(2, ws.max_row + 1):
                name = ws.cell(r, col_name).value
                if not name or not str(name).strip():
                    skipped += 1
                    continue
                name = str(name).strip()

                # region_id: xlsx 给的是 regions.code,需要转换为 regions.id(cuid)
                raw_region_code = (
                    ws.cell(r, col_region_id).value if col_region_id else None
                )
                region_id = None
                if raw_region_code:
                    code = str(raw_region_code).strip()
                    region_id = code_to_id.get(code)
                    if region_id is None:
                        region_unmatched += 1

                # alias: JSON 字符串 → list
                alias_raw = ws.cell(r, col_alias).value if col_alias else None
                alias_val = parse_json_cell(alias_raw)
                alias_json = (
                    json.dumps(alias_val, ensure_ascii=False)
                    if alias_val is not None
                    else None
                )

                # address
                addr_raw = ws.cell(r, col_address).value if col_address else None
                addr_val = parse_json_cell(addr_raw)
                addr_json = (
                    json.dumps(addr_val, ensure_ascii=False)
                    if addr_val is not None
                    else None
                )

                # subarea: 当前 DB 没该列,先不写(预留读取,后续 DDL 加回后启用)
                subarea_raw = ws.cell(r, col_subarea).value if col_subarea else None
                subarea_val = parse_json_cell(subarea_raw)
                subarea_json = (
                    json.dumps(subarea_val, ensure_ascii=False)
                    if subarea_val is not None
                    else None
                )

                # 写库。subarea 列当前不存在 → 暂不写入,等 DDL 加回后补字段
                cur.execute(
                    """
                    INSERT INTO communities (
                        id, name, alias, region_id, address, status, created_at, updated_at
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, NOW(), NOW()
                    )
                    """,
                    (
                        _gen_cuid(),
                        name,
                        alias_json,
                        region_id,
                        addr_json,
                        1,
                    ),
                )
                # TODO(subarea):当 DDL 加回 subarea 列后,把 subarea_json 加到 INSERT 中
                inserted += 1

        conn.commit()
        print(
            f"\n导入完成: 新增 {inserted} 条 / 跳过 {skipped} 条 (空名) "
            f"/ region 未匹配 {region_unmatched} 条"
        )
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ————————————————————— cuid 生成 —————————————————————

import secrets
import time


def _gen_cuid() -> str:
    """生成一个 24 字符的伪 cuid(cuid v1 简化版)。

    真实 cuid 含 timestamp + counter + fingerprint + random。本脚本不追求格式严格,
    只保证 24 位、以字母开头、含字母数字,适配 `varchar(191) NOT NULL`。
    """
    ts = format(int(time.time() * 1000), "016x")  # 16 chars hex
    rnd = secrets.token_hex(4)  # 8 chars hex
    return f"c{ts}{rnd}"  # 25 chars,以 c 开头


if __name__ == "__main__":
    import_communities()
