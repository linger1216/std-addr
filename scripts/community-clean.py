#!/usr/bin/env python3
"""
清洗《居村委.xlsx》里的"居委"数据,输出可核对/可导入 communities 表的新 excel。

只处理 类型=='居委' 的行(村委不是小区,跳过)。

已知处理规则(对应用户给的 7 条名称规则 + 4 条地址规则):
  1. "一品漫城一、二期" 这类"数字/中文数字、数字/中文数字 + 期"的后缀 —— 拆成多条记录,
     每条共用原行地址。
  2. "万特园(臻园)" —— 括号内容当别名,只在"清洗后名称+地址"组合只出现一行时生效。
  3. "上海康城(一期)" —— 括号内容是纯"N期"标记,合并进名称,不拆分、不当别名。
  4. "上虹新村1-6街坊"/"上虹新村七-九街坊" —— 硬编码展开成"一期".."九期"(见
     _HARDCODED_NAME_EXPANSIONS)。这个模式在源数据里还有 50+ 种"N街坊"/"N-M号"变体,
     本次只按用户给的这一个例子处理,其余不动、标记待复核,避免自动展开出错。
  5. "东方御花园(东方御花园(北区))" —— 括号内容以外层名称自身开头(自我重复嵌套),
     去重合并成"东方御花园北区"。
  6. "中虹浦江苑(浦涛路100弄)" —— 括号内容是地址,不是别名,整体丢弃。
  7. "合生城邦二街坊(剑桥府邸)"/"(南洋瑞都)"/"(金榜雅苑)" —— 同名同地址出现多行,
     括号内容合并进 subarea 数组,3 行收敛成 1 条记录。

  别名 vs 子区域的判断规则(用户确认):"清洗后名称+地址"这个组合只出现 1 行 → 括号
  内容当别名(alias,数组里通常 1 个元素);出现 >=2 行 → 括号内容当子区域
  (subarea,数组里通常多个元素),多行合并成 1 条记录。

地址清洗规则(**已弃用** —— 见下):
  - 去掉《...》书名号包住的内容(小区曾用名/别称类噪音)
  - 去掉"N户"这种入户数噪音
  - 去掉(...)/(...)括号内容(要么是地址补充说明,要么是无效值,统一按用户给的例子丢弃)
  - 剩余文本按 、/，/,/;/； 切分成多个地址

⚠️ 上述地址清洗只在「fallback 路径」 使用。地址优先由 scripts/community-address.txt 提供:
  - 文件每行: <序号>\t<学校>\t<校址>\t<招生范围>
  - 解析时按 [、,;,;] 切段(括号内不切),每段第一个外层括号内的内容作为该小区的地址列表
  - 段前缀作为小区候选名,纯路名段(无括号)直接丢弃
  - 清洗后小区名 vs 文件小区名做 difflib.SequenceMatcher.ratio 匹配,产出
    「候选小区名 / 候选地址 / 来源学校 / 匹配度」 4 列供人工核对。
  - 「地址」 列采用规则:
      1. 名称含「期」 → 旧清洗(教育文件不区分期)
      2. 名称不含「期」 且 教育文件「唯一完美命中」(score == 1.0 且并列冠军 == 1) → 教育文件
      3. 其他情况 → fallback 到旧清洗(依赖本身的地址处理)

所属居村委 → region_id:
  从 scripts/region.json 递归读出所有居/村委会叶子节点(每个节点带 12 位
  addressStandardCode 作为 region_id,以及所属街镇)。脚本按"名称" 剥掉
  "XX居民委员会"/"XX居委会"/"XX村委会"/"XX村民委员会"等后缀,取核心词做
  精确匹配;精确匹配不到时用 difflib 模糊匹配(阈值 0.6),仍匹配不到就保留
  该行、region_id 留空、在"待复核"列写明原因 —— 不丢数据,交给人工在 excel 里核对。

用法:
    python3 scripts/community-clean.py
输出:
    community_cleaned.xlsx (仓库根目录)
依赖(需要先装):
    pip3 install openpyxl
"""

from __future__ import annotations

import difflib
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC_XLSX = ROOT / "居村委.xlsx"
SRC_ADDR_TXT = ROOT / "scripts" / "community-address.txt"
SRC_REGION_JSON = ROOT / "scripts" / "region.json"
OUT_XLSX = ROOT / "community_cleaned.xlsx"

# ————————————————————— 1. regions 参考表 —————————————————————

_COMMITTEE_SUFFIXES = [
    "居民委员会",
    "村民委员会",
    "居委会",
    "村委会",
    "居委",
    "村委",
]

# JSON 里识别居/村委会节点的关键词。包含「居民/村民委员会」全称,因为 JSON 里
# 大部分节点是「某某居民委员会」「某某村民委员会」(中间有「民」),而 DB 旧数据
# 多是「某某居委会」。同一份 `strip_committee_suffix` 处理两种形态。
# 注意:
#   - 「委员会」 太宽(「财政经济委员会」 等区级专委会会误中),不放进来
#   - 「居委/村委」 是「居委会/村委会」 子串,但「农业与农村委员会」 也含「村委」,
#     所以用 _is_committee_node 时按边界判断,严格命中名字末尾的「居委/村委」,
#     避免「农村委员会」 误中
_COMMITTEE_NODE_KEYWORDS = (
    "居民委员会",
    "村民委员会",
    "居委会",
    "村委会",
)


def strip_committee_suffix(name: str) -> str:
    """'一品漫城第一居委会' -> '一品漫城第一';'华漕第二居民委员会' -> '华漕第二'"""
    name = (name or "").strip()
    for suf in _COMMITTEE_SUFFIXES:
        if name.endswith(suf):
            return name[: -len(suf)]
    return name


def _is_committee_node(name: str) -> bool:
    """名字命中居/村委会关键词,且「居委/村委」 这类短词必须在末尾(避免「农村委员会」 误中)。"""
    for k in _COMMITTEE_NODE_KEYWORDS:
        if k in name:
            return True
    # 「居委/村委」 短词必须在末尾才算
    return name.endswith("居委") or name.endswith("村委")


def _walk_region_tree(node: dict, parents: list[str], town_hint: str | None):
    """递归遍历 region.json 树。yield (code, name, town) 三元组。

    节点判定:
      - 节点含居委会/村委会类关键词,且 childList 为空 → 叶子,产出
      - 否则继续递归子节点
    town_hint 沿父链上最近的「镇/街道」 名字。
    """
    name = node.get("orgName", "") or ""
    code = node.get("addressStandardCode", "") or ""
    children = node.get("childList") or []

    # 镇上「镇/街道」 节点名字
    if ("镇" in name and len(name) <= 6) or "街道" in name:
        town_hint = name

    if not children and _is_committee_node(name):
        yield (code, name, town_hint or "")

    for c in children:
        yield from _walk_region_tree(c, parents + [name], town_hint)


def load_region_committees(path: Path = SRC_REGION_JSON) -> list[tuple[str, str, str]]:
    """从 scripts/region.json 读居委会/村委会列表。

    返回 [(code, name, town), ...]
      - code: 12 位 addressStandardCode(原 DB 的 id 字段)
      - name: 居/村委会全名(「景华新苑居民委员会」 等)
      - town: 所属街镇(从父链推断;顶层无镇时为空串)

    跳过 code 为空的节点(筹备组等未落码的)。
    """
    if not path.exists():
        raise FileNotFoundError(f"找不到 region.json: {path}")
    payload = json.loads(path.read_text(encoding="utf-8"))
    # 容错:接受裸 list 或带 envelope 的 {data: [...]} 结构
    if isinstance(payload, dict) and "data" in payload:
        root_list = payload["data"]
    elif isinstance(payload, list):
        root_list = payload
    else:
        raise RuntimeError(f"region.json 结构无法识别: 顶层 {type(payload).__name__}")

    out: list[tuple[str, str, str]] = []
    for top in root_list:
        # 顶层 orgName 一般是「闵行区」,跳过避免污染 town
        for code, name, town in _walk_region_tree(top, [top.get("orgName", "")], None):
            if not code:
                continue  # 筹备组等无编码的节点丢弃
            out.append((code, name, town))
    return out


class RegionMatcher:
    """所属居村委原始值 -> region.json addressStandardCode 的匹配器。"""

    FUZZY_THRESHOLD = 0.6

    def __init__(self, committees: list[tuple[str, str, str]]):
        """committees: [(code, name, town), ...] —— 来自 load_region_committees()"""
        self._by_core: dict[str, tuple[str, str, str]] = {}
        self._entries: list[tuple[str, str, str, str]] = []
        for code, name, town in committees:
            core = strip_committee_suffix(name)
            self._entries.append((code, name, town, core))
            # 同核心词多个 region 时保留先出现的一个,冲突不在本脚本里解
            self._by_core.setdefault(core, (code, name, town))

    def match(self, raw: str) -> tuple[str | None, str | None, float, str]:
        """返回 (region_id, region_name, score, note)"""
        core = strip_committee_suffix(raw)
        exact = self._by_core.get(core)
        if exact:
            # region_id 用 code,region_name 用全名(含「居委会/居民委员会」 后缀)
            return exact[0], exact[1], 1.0, ""

        best: tuple[str, str, str] | None = None
        best_score = 0.0
        for code, name, _town, ecore in self._entries:
            score = difflib.SequenceMatcher(None, core, ecore).ratio()
            if score > best_score:
                best_score = score
                best = (code, name, _town)

        if best and best_score >= self.FUZZY_THRESHOLD:
            return best[0], best[1], round(best_score, 2), f"模糊匹配(相似度{best_score:.2f})"
        return None, None, round(best_score, 2), f"未匹配居委「{raw}」"


# ————————————————————— 2. 地址清洗 —————————————————————

_BOOK_MARK_RE = re.compile(r"《[^》]*》")
_HOUSEHOLD_COUNT_RE = re.compile(r"\d+户")
_FULLWIDTH_PAREN_RE = re.compile(r"（[^（）]*）")
_ASCII_PAREN_RE = re.compile(r"\([^()]*\)")
# 段落分隔:「，,;；」 在地址里通常是「另一条独立地址」的边界。
# 例:「富岩路155弄、156弄，昆阳路2038弄」 用「，」 把两个路名分开。
# 注意:这里 **不包含「、」**,因为段内还会按「、」 进一步展开(见 _SAME_TAIL_RE)。
_ADDR_SEGMENT_SPLIT_RE = re.compile(r"[，,;；]")
# 段内 "n、m 共享同一后缀" 模式:
#   前缀(路名/小区名,可含数字) + 数字列表「n、m、k」 + 同一后缀(弄/号/支弄 + 可选数字)。
#   例:
#     「富岩路155弄、156弄」       -> 「富岩路155弄」/「富岩路156弄」
#     「七莘路3128弄38、25支弄」 -> 「七莘路3128弄38支弄」/「七莘路3128弄25支弄」
#   后缀里必须出现「路/弄/号/支弄/街/道/巷/村/组」 中至少一个路名要素,避免误拆普通列表。
_CJK_OR_DIGIT = r"[一二三四五六七八九十0-9]+"
_SAME_TAIL_RE = re.compile(
    rf"^(?P<head>.*?)(?P<nums>{_CJK_OR_DIGIT}(?:[、]{_CJK_OR_DIGIT})+)(?P<tail>(?:支弄|弄|号|路|街|道|巷|村|组).*)$"
)
# 兜底:如果段内只出现一次「、」 切完是两个完整地址(prefix + nums + tail + tail),不算
# "n、m 共享后缀" 模式。这种就当两条独立地址。
_DUNHAO_SPLIT_RE = re.compile(r"、")


def _expand_same_tail(segment: str) -> list[str] | None:
    """段内 "n、m 共享同一后缀" 展开。

    例: '富岩路155弄、156弄' -> ['富岩路155弄','富岩路156弄']
        '七莘路3128弄38、25支弄' -> ['七莘路3128弄38支弄','七莘路3128弄25支弄']
    不匹配返回 None。
    """
    m = _SAME_TAIL_RE.match(segment)
    if not m:
        return None
    head = m.group("head")
    nums = re.split("、", m.group("nums"))
    tail = m.group("tail")
    return [f"{head}{n}{tail}" for n in nums]


def clean_addresses(raw: str | None) -> list[str]:
    """地址清洗。

    规则:
      1. 先剥书名号 / 户数 / 括号噪音。
      2. 按「，,;；」 把原文切成多段(每段是一条独立地址)。
      3. 每段尝试 "n、m 共享后缀" 展开(例 「富岩路155弄、156弄」 -> 两条)。
         不匹配则段内「、」 视为地址边界,按「、」 切。

    返回去空/去 strip 后的地址列表。
    """
    if not raw:
        return []
    s = raw
    s = _BOOK_MARK_RE.sub("", s)
    s = _HOUSEHOLD_COUNT_RE.sub("", s)
    s = _FULLWIDTH_PAREN_RE.sub("", s)
    s = _ASCII_PAREN_RE.sub("", s)

    result: list[str] = []
    for segment in _ADDR_SEGMENT_SPLIT_RE.split(s):
        segment = segment.strip()
        if not segment:
            continue
        expanded = _expand_same_tail(segment)
        if expanded is not None:
            for e in expanded:
                e = e.strip()
                if e:
                    result.append(e)
            continue
        # 兜底:段内只出现「、」,按字面切。
        if "、" in segment:
            for e in _DUNHAO_SPLIT_RE.split(segment):
                e = e.strip()
                if e:
                    result.append(e)
        else:
            result.append(segment)
    return result


# ————————————————————— 3. 名称清洗 —————————————————————

# 规则 4:上虹新村街坊范围 -> 期。数据里同类"N街坊"/"N-M街坊"还有 50+ 种变体,
# 本次只按用户给的例子硬编码这一组,其余保持原样 + 标记待复核。
_HARDCODED_NAME_EXPANSIONS: dict[str, list[str]] = {
    "上虹新村1-6街坊": [f"上虹新村{n}期" for n in "一二三四五六"],
    "上虹新村七-九街坊": [f"上虹新村{n}期" for n in "七八九"],
}

# 规则 1:"一品漫城一、二期" / 括号内 "（一、二期）" / "（一期、二期、三期）" 都按期列表展开。
# 接受两种串形态:
#   1. 尾期形态:nums[、,]nums期,例 "一、二期"、"1、2期" → nums="一、二", 尾巴="期"
#   2. 每项带期:每项都带「期」,例 "一期、二期"、"一期、二期、三期" → 按「、」切,每项就是一个小
#      区名,不再拼「期」
_PERIOD_TAIL_FORM_RE = re.compile(rf"^(?P<nums>(?:{_CJK_OR_DIGIT}[、，])+{_CJK_OR_DIGIT})期$")
# 切「一期、二期、三期」时,每项独立校验:含「期」 且前面部分是数字/中文数字。
_PERIOD_ITEM_RE = re.compile(rf"^{_CJK_OR_DIGIT}期$")
# 规则 3:纯 "N期" 标记(不含顿号/逗号)
_PERIOD_ONLY_RE = re.compile(rf"^{_CJK_OR_DIGIT}期$")
# 括号内容长得像地址(规则 6):含"路/弄/号/街/道/巷"紧跟数字,或书名号,或纯数字-数字
_ADDRESS_LIKE_RE = re.compile(r"[路弄号街道巷组]\d|《|^\d+[-~]\d+")
# 清洗后名称里如果还留着这些符号,大概率是没吃透的一次性花样,标记待复核
_LEFTOVER_JUNK_RE = re.compile(r"[《》()]")
# 含「期」 的 pending_bracket 不能进 alias/subarea —— 这本该是小区名(规则 3)而不是别名。
# 正常路径下 N期/N、M期 已经在 _normalize_name_inner 里展开成 name,不会到这一层;
# 这里做防御:万一上游漏掉,也不要让「期」 串进别名列。
_PENDING_HAS_PERIOD_RE = re.compile(r"期")
# 同一条规则也用于:决定该条小区记录的「地址」列用哪种来源
# 含「期」→ 旧清洗(clean_addresses,基于源 xlsx「地址」字段)
# 不含「期」→ 教育文件(community-address.txt 招生范围)
_NAME_HAS_PERIOD_RE = _PENDING_HAS_PERIOD_RE
# 名称里还留着"N-M"这种数字/中文数字范围(街坊/号/弄等),没有命中硬编码展开表,
# 大概率是源数据里另一种"按范围分小区"的写法,本次不猜,标记待复核
_UNEXPANDED_RANGE_RE = re.compile(
    r"[0-9]+[-~][0-9]+|[一二三四五六七八九十]+[-~][一二三四五六七八九十]+"
)


def split_period_list(suffix: str) -> list[str] | None:
    """期列表拆开。

    支持两种形态:
      1. 尾期形态 '一、二期' -> ['一期','二期'](用户给定的早期用例)
      2. 每项带期 '一期、二期、三期' -> ['一期','二期','三期'](当前用户给的形态)

    不匹配返回 None。
    """
    # 形态 1:尾期
    m = _PERIOD_TAIL_FORM_RE.match(suffix)
    if m:
        nums = re.split("[、，]", m.group("nums"))
        return [f"{n}期" for n in nums]
    # 形态 2:每项带期。按「、」/「，」切,每项都得是 N期 形式,否则不认。
    if "、" in suffix or "，" in suffix:
        parts = re.split("[、，]", suffix)
        if all(_PERIOD_ITEM_RE.match(p.strip()) for p in parts):
            return [p.strip() for p in parts]
    return None


def split_outer_paren(s: str) -> tuple[str, str | None, str]:
    """找最外层一组括号(全角「（）」 或 半角「()」),返回 (括号前, 括号内容 or None, 括号后)。"""
    # 优先全角,找不到再试半角(避免「(全角+半角混用」 的边界怪例)
    for open_ch, close_ch in (("（", "）"), ("(", ")")):
        i = s.find(open_ch)
        if i == -1:
            continue
        depth = 0
        for j in range(i, len(s)):
            if s[j] == open_ch:
                depth += 1
            elif s[j] == close_ch:
                depth -= 1
                if depth == 0:
                    return s[:i], s[i + 1 : j], s[j + 1 :]
        return s, None, ""  # 同种括号不闭合,不再尝试另一种,直接当没括号
    return s, None, ""


class NamePiece:
    """一次名称清洗产出的一小片:后面按 (name, committee) 分组决定 alias/subarea。"""

    __slots__ = ("name", "pending_bracket", "review_notes")

    def __init__(self, name: str, pending_bracket: str | None, review_notes: list[str]):
        self.name = name
        self.pending_bracket = pending_bracket
        self.review_notes = review_notes


def normalize_name(raw: str) -> list[NamePiece]:
    raw = (raw or "").strip()

    # 规则 4:硬编码展开(已知形状,不需要范围校验)
    if raw in _HARDCODED_NAME_EXPANSIONS:
        return [NamePiece(n, None, []) for n in _HARDCODED_NAME_EXPANSIONS[raw]]

    pieces = _normalize_name_inner(raw)
    for piece in pieces:
        if _UNEXPANDED_RANGE_RE.search(piece.name):
            piece.review_notes.append(
                f"名称含数字/中文数字范围,未自动展开(仅上虹新村按硬编码处理): {piece.name}"
            )
    return pieces


def _normalize_name_inner(raw: str) -> list[NamePiece]:
    # 规则 1:名称整体就是 "底 + N、M期" 后缀(不带括号)
    # 注意:半角括号和全角括号都视作括号,不能跳过 split_outer_paren 路径
    if "（" not in raw and "(" not in raw:
        for cut in range(len(raw) - 1, -1, -1):
            base, suffix = raw[:cut], raw[cut:]
            nums = split_period_list(suffix)
            if nums:
                return [NamePiece(base + n, None, []) for n in nums]
        return [NamePiece(raw, None, [])]

    before, content, after = split_outer_paren(raw)
    if content is None:
        # split_outer_paren 同时认全角/半角,都没匹配上时(只有单边括号等)原样返回
        return [NamePiece(raw, None, [])]

    before = before.strip()
    content = content.strip()
    after = after.strip()
    if after:
        # 括号后面还有尾巴,不是本次覆盖的形状,原样保留 + 标记
        return [NamePiece(raw, None, [f"名称结构未识别(括号后有尾缀): {raw}"])]

    # 规则 5:嵌套自我重复,如 "东方御花园(东方御花园(北区))"
    if content.startswith(before) and content != before:
        inner = content[len(before) :].strip()
        inner_before, inner_content, inner_after = split_outer_paren(inner)
        if inner_content is not None and not inner_before and not inner_after:
            inner = inner_content.strip()
        merged = before + inner
        notes = [f"名称结构未识别(去重后仍有括号): {raw}"] if _LEFTOVER_JUNK_RE.search(merged) else []
        return [NamePiece(merged, None, notes)]

    # 规则 1(括号版):"复地北桥城(一、二期)"
    period_list = split_period_list(content)
    if period_list:
        return [NamePiece(before + n, None, []) for n in period_list]

    # 规则 3:纯 "N期" 标记,合并不拆分
    if _PERIOD_ONLY_RE.match(content):
        return [NamePiece(before + content, None, [])]

    # 规则 6:括号内容长得像地址,整体丢弃
    if _ADDRESS_LIKE_RE.search(content):
        return [NamePiece(before, None, [])]

    # 规则 2/7:候选别名或子区域,留到分组阶段按"同名同址出现次数"判断
    return [NamePiece(before, content, [])]


# ————————————————————— 3.5 招生范围地址文件解析 —————————————————————


# 段拆分:不在括号内的 [、，,;；] 都算分段。
# 注意:全角/半角都要识别,「、【】」 紧贴括号内文时不能切。
_SEGMENT_SPLIT_RE = re.compile(r"[、，,;；]")
# 括号内再切:全角顿号 + 半角逗号;冒号不考虑,因为文件里实际没有。
_ADDR_INSIDE_SPLIT_RE = re.compile(r"[、，,]")
_PAREN_OPEN_RE = re.compile(r"[（(]")
_PAREN_CLOSE_RE = re.compile(r"[）)]")


def _split_segments_outside_paren(text: str) -> list[str]:
    """把招生范围按「、，,;；」 切段,但括号内部不切。

    实现:逐字符扫描,维护 depth 计数(全角/半角括号各算一层)。
    """
    segments: list[str] = []
    buf: list[str] = []
    depth = 0
    for ch in text:
        if ch in "（(":
            depth += 1
            buf.append(ch)
        elif ch in "）)":
            depth = max(0, depth - 1)
            buf.append(ch)
        elif depth == 0 and _SEGMENT_SPLIT_RE.match(ch):
            seg = "".join(buf).strip()
            if seg:
                segments.append(seg)
            buf = []
        else:
            buf.append(ch)
    tail = "".join(buf).strip()
    if tail:
        segments.append(tail)
    return segments


def _find_first_outer_paren(s: str) -> tuple[str, str | None, str] | None:
    """找第一组最外层括号,返回 (前, 内容, 后);没有括号返回 None。

    全角和半角括号都接受。优先匹配第一个出现的(无论全/半角)。
    """
    # 找最早出现的左括号(全角或半角)
    candidates = [(m.start(), m.group()) for m in _PAREN_OPEN_RE.finditer(s)]
    if not candidates:
        return None
    start, open_ch = candidates[0]
    close_ch = "）)"[ "（(".index(open_ch)]
    depth = 0
    for j in range(start, len(s)):
        if s[j] == open_ch:
            depth += 1
        elif s[j] == close_ch:
            depth -= 1
            if depth == 0:
                return s[:start], s[start + 1 : j], s[j + 1 :]
    return None  # 括号不闭合


class AddressEntry:
    """招生范围文件里的一条 (小区, 地址列表, 来源学校)。"""

    __slots__ = ("name", "addresses", "school")

    def __init__(self, name: str, addresses: list[str], school: str):
        self.name = name
        self.addresses = addresses
        self.school = school


def parse_address_file(path: Path = SRC_ADDR_TXT) -> list[AddressEntry]:
    """解析 community-address.txt,返回所有 (小区, 地址列表, 学校) 列表。

    规则:
      - 文件每行: <序号>\\t<学校>\\t<校址>\\t<招生范围>
      - 招生范围按 [、,;,;] 切段(括号内不切)
      - 段有第一个外层括号 -> 括号内容按 [、,,] 切成地址数组,前缀作为小区名
      - 段没有括号 -> 丢弃(纯路名段如「汇通路」 不是完整小区)
      - 段有多个括号 -> 只取第一个作为地址;其余内容记到 review(本版本丢弃)
    """
    if not path.exists():
        raise FileNotFoundError(f"找不到招生范围文件: {path}")

    entries: list[AddressEntry] = []
    for lineno, raw in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        line = raw.strip()
        if not line or line.startswith("2026年") or lineno <= 2:
            # 第 1 行是标题,第 2 行是表头
            continue
        # tab 或多个空格分隔
        parts = re.split(r"\t+|\s{2,}", line)
        if len(parts) < 4:
            # tab 至少要有 3 个才能拿到范围列
            continue
        # parts[0] 序号, parts[1] 学校, parts[2] 校址, parts[3..] 范围(可能含 tab)
        school = parts[1].strip()
        # 重新拼回去,避免 tab 把范围内多空格搞乱
        # 文件每行固定 3 个 tab:序号 \t 学校 \t 校址 \t 范围
        first_cut = line.find("\t")
        second_cut = line.find("\t", first_cut + 1)
        third_cut = line.find("\t", second_cut + 1)
        if third_cut == -1:
            # 没有第 4 列(有些行是「详见学校招生简章」)
            scope = ""
        else:
            scope = line[third_cut + 1 :].strip()
        if not scope or scope == "详见学校招生简章":
            continue

        for seg in _split_segments_outside_paren(scope):
            seg = seg.strip().strip("、,，;；").strip()
            if not seg:
                continue
            paren = _find_first_outer_paren(seg)
            if paren is None:
                # 没括号 -> 纯路名/区域段,丢弃(不是完整小区)
                continue
            before, content, _after = paren
            name = before.strip().strip("、,，;；").strip()
            # 去掉常见前缀噪音:「【以下小区为暂时安置】」 等
            name = re.sub(r"^【[^】]*】", "", name).strip()
            content = content.strip()
            if not name or not content:
                continue
            addresses = [a.strip() for a in _ADDR_INSIDE_SPLIT_RE.split(content) if a.strip()]
            if not addresses:
                continue
            entries.append(AddressEntry(name=name, addresses=addresses, school=school))
    return entries


def best_match_score(
    cleaned_name: str,
    addr_entries: list[AddressEntry],
) -> tuple[AddressEntry | None, float, int]:
    """在 addr_entries 里找跟 cleaned_name 匹配度最高的条目。

    返回 (best_entry, best_score, ties_count):
      - best_entry: 得分最高的条目(并列冠军时取第一个)
      - best_score: 最高得分(已四舍五入)
      - ties_count: 同样达到 best_score 的条目数(并列冠军数,含 best_entry)

    score 用 difflib.SequenceMatcher.ratio;若 cleaned_name 命中某个 entry.name 的子串
    (小区名是子串如 「春城」 命中 「上海春城」),给一定加分(粗略启发式)。

    业务侧通过 ties_count 判断「唯一命中」:
      - ties_count == 1 且 best_score == 1.0 → 唯一完美命中,可用其地址
      - ties_count >= 2 → 多条并列,需人工核对
      - best_score < 1.0 → 模糊匹配,需要人工核对
    """
    if not addr_entries:
        return None, 0.0, 0
    cn = cleaned_name.strip()
    scored: list[tuple[AddressEntry, float]] = []
    for e in addr_entries:
        ratio = difflib.SequenceMatcher(None, cn, e.name).ratio()
        # 子串加分:cn 是 entry.name 子串 或反之,且长度差不悬殊(>= 1/2)
        # 上限 1.0 —— 业务侧用 `score == 1.0` 判断唯一完美命中,超出 1.0 反而
        # 走 fallback,失去加分意义。
        bonus = 0.0
        if cn and e.name:
            shorter, longer = (cn, e.name) if len(cn) <= len(e.name) else (e.name, cn)
            if shorter in longer and len(shorter) / len(longer) >= 0.5:
                bonus = 0.15
        score = round(min(ratio + bonus, 1.0), 2)
        scored.append((e, score))

    best_score = max(s for _, s in scored)
    if best_score <= 0:
        return None, 0.0, 0
    # 并列冠军:同分的所有 entry
    winners = [e for e, s in scored if s == best_score]
    return winners[0], best_score, len(winners)


# ————————————————————— 4. 主流程 —————————————————————


class Row:
    __slots__ = (
        "name",
        "committee_raw",
        "addresses",
        "source_row",
        "review_notes",
        "pending_bracket",
        # 源 excel 4 列原始值,保留下来以便核对/回溯(展开/拆分后原值仍需可见)
        "name_raw",
        "type_raw",
        "address_raw",
    )

    def __init__(
        self,
        name,
        committee_raw,
        addresses,
        source_row,
        review_notes,
        pending_bracket,
        name_raw,
        type_raw,
        address_raw,
    ):
        self.name = name
        self.committee_raw = committee_raw
        self.addresses = addresses
        self.source_row = source_row
        self.review_notes = review_notes
        self.pending_bracket = pending_bracket
        self.name_raw = name_raw
        self.type_raw = type_raw
        self.address_raw = address_raw


def load_source_rows() -> list[tuple]:
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    assert header[:4] == ("名称", "所属居村委", "类型", "地址"), f"表头不符合预期: {header}"
    return rows[1:]


def expand_rows(raw_rows: list[tuple]) -> list[Row]:
    expanded: list[Row] = []
    for idx, (name, committee, type_, address) in enumerate(raw_rows, start=2):  # excel 行号(含表头)
        if (type_ or "").strip() != "居委":
            continue
        addresses = clean_addresses(address)
        pieces = normalize_name(name)
        for piece in pieces:
            notes = list(piece.review_notes)
            if not addresses:
                notes.append("地址清洗后为空")
            expanded.append(
                Row(
                    name=piece.name,
                    committee_raw=(committee or "").strip(),
                    addresses=addresses,
                    source_row=idx,
                    review_notes=notes,
                    pending_bracket=piece.pending_bracket,
                    name_raw=(name or "").strip(),
                    type_raw=(type_ or "").strip(),
                    address_raw=(address or "").strip(),
                )
            )
    return expanded


class Group:
    __slots__ = (
        "name",
        "committee_raw",
        "addresses",
        "alias",
        "subarea",
        "source_rows",
        "review_notes",
        # 同组对应的源 excel 原始值(多条原始行合成一条 Group 时,保留首条,
        # 同组里如果出现不同原始值,在 review_notes 里追加说明)
        "name_raw",
        "type_raw",
        "address_raw",
    )

    def __init__(self, name: str, committee_raw: str):
        self.name = name
        self.committee_raw = committee_raw
        self.addresses: list[str] = []
        self.alias: list[str] = []
        self.subarea: list[str] = []
        self.source_rows: list[int] = []
        self.review_notes: list[str] = []
        self.name_raw: str = ""
        self.type_raw: str = ""
        self.address_raw: str = ""

    def add(self, row: Row):
        for a in row.addresses:
            if a not in self.addresses:
                self.addresses.append(a)
        self.source_rows.append(row.source_row)
        self.review_notes.extend(row.review_notes)
        # 首条记录到的原始值作为该组的原始值,后续行只在不一致时打 review_note
        if not self.name_raw:
            self.name_raw = row.name_raw
            self.type_raw = row.type_raw
            self.address_raw = row.address_raw
        else:
            if row.name_raw and row.name_raw != self.name_raw:
                self.review_notes.append(
                    f"第{row.source_row}行原始名称「{row.name_raw}」与同组首行不一致"
                )
            if row.address_raw and row.address_raw != self.address_raw:
                self.review_notes.append(
                    f"第{row.source_row}行原始地址「{row.address_raw}」与同组首行不一致"
                )
        if row.committee_raw and row.committee_raw != self.committee_raw:
            self.review_notes.append(
                f"第{row.source_row}行所属居村委「{row.committee_raw}」与同组不一致"
            )


def group_rows(rows: list[Row]) -> list[Group]:
    order: list[tuple[str, str]] = []
    groups: dict[tuple[str, str], Group] = {}
    pending_by_group: dict[tuple[str, str], list[str]] = {}

    for row in rows:
        key = (row.name, row.committee_raw)
        if key not in groups:
            groups[key] = Group(row.name, row.committee_raw)
            pending_by_group[key] = []
            order.append(key)
        groups[key].add(row)
        if row.pending_bracket:
            pending_by_group[key].append(row.pending_bracket)

    result = []
    for key in order:
        g = groups[key]
        pendings = pending_by_group[key]
        # 「期」 串不能进 alias/subarea(规则 3/用户确认)。
        # 正常路径 N期/N、M期 已经在 _normalize_name_inner 里展开成多条 name,
        # 走不到 pending_bracket;这里再过滤一次是防御性兜底。
        safe_pendings = [p for p in pendings if not _PENDING_HAS_PERIOD_RE.search(p)]
        dropped = [p for p in pendings if p not in safe_pendings]
        if dropped:
            g.review_notes.append(
                f"含「期」的括号内容按规则 3 应是小区名而非别名,本组未自动展开: {dropped}"
            )
        # 用户确认的判断规则:同名同址只出现 1 行 -> 别名;出现多行 -> 子区域
        if len(g.source_rows) == 1:
            if safe_pendings:
                g.alias = [safe_pendings[0]]
        else:
            seen = []
            for p in safe_pendings:
                if p and p not in seen:
                    seen.append(p)
            g.subarea = seen
        result.append(g)
    return result


def build_output(
    groups: list[Group],
    matcher: RegionMatcher,
    addr_entries: list[AddressEntry] | None = None,
) -> None:
    """生成 community_cleaned.xlsx。

    主 sheet「communities」:
      - 原有清洗列(名称/原始字段/别名/区划匹配/子区域/状态/来源行号)
      - 「地址」 列来源决策表:
          1. 名称含「一期/二期/三期」 等 → 旧清洗(clean_addresses,源 xlsx)
          2. 名称不含「期」 且 教育文件里「唯一完美命中」
             (score == 1.0 且并列冠军数 == 1)→ 教育文件
          3. 其他(无候选 / score < 1.0 / 多条并列)→ fallback 到旧清洗
        「地址来源」 列明示这一选择,「待复核」 列附 fallback 原因。
      - 末尾追加 4 列(候选小区名/候选地址列表/来源学校/匹配度) —— 教育文件匹配结果,
        不论地址列采用哪种策略,这几列都给,供人工核对。

    备选 sheet「address_src」:
      - 文件里所有 (小区, 地址, 学校) 平铺,供用户直接查表。
    """
    addr_entries = addr_entries or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "communities"
    ws.append(
        [
            "名称",
            "原始名称",
            "原始所属居村委",
            "原始类型",
            "原始地址",
            "别名",
            "匹配区划名称",
            "区划ID",
            "匹配置信度",
            "子区域",
            "状态",
            "地址",
            "地址来源",
            "候选小区名(文件)",
            "候选地址(文件)",
            "来源学校",
            "匹配度",
            "待复核",
            "来源excel行号",
        ]
    )

    matched, unmatched, flagged = 0, 0, 0
    for g in groups:
        region_id, region_name, score, match_note = matcher.match(g.committee_raw)
        if region_id:
            matched += 1
        else:
            unmatched += 1

        # 在文件里找最佳候选(带并列冠军数)
        best_entry, best_score, ties_count = best_match_score(g.name, addr_entries)
        if best_entry is not None:
            cand_name = best_entry.name
            cand_addrs = json.dumps(best_entry.addresses, ensure_ascii=False)
            cand_school = best_entry.school
            cand_score = best_score
        else:
            cand_name = cand_addrs = cand_school = cand_score = None

        # 「地址」 列来源决策:
        #   1. 名字含「期」 → 用旧清洗(教育文件不区分期)
        #   2. 名字不含「期」 且 教育文件里「唯一完美命中」(score == 1.0 且并列数 == 1)→ 用教育文件
        #   3. 其他 → fallback 到旧清洗
        has_period = bool(_NAME_HAS_PERIOD_RE.search(g.name))
        unique_perfect = (
            best_entry is not None
            and best_score == 1.0
            and ties_count == 1
        )
        if has_period:
            chosen_addrs = list(g.addresses)
            addr_source = "旧清洗(源xlsx)"
            addr_note = None  # 不需要额外说明,名字已含「期」
        elif unique_perfect:
            chosen_addrs = list(best_entry.addresses)
            addr_source = "教育文件"
            addr_note = None  # 唯一完美命中,无需说明
        else:
            chosen_addrs = list(g.addresses)
            addr_source = "旧清洗(源xlsx)"
            if best_entry is None:
                addr_note = "教育文件未找到候选,fallback 到源xlsx地址"
            elif best_score < 1.0:
                addr_note = f"教育文件无完美命中(最高 {best_score}),fallback 到源xlsx地址"
            else:
                # ties_count >= 2
                addr_note = f"教育文件并列 {ties_count} 条候选(均 score={best_score}),fallback 到源xlsx地址"
        if not chosen_addrs:
            addr_source = "无"
            addr_note = addr_note or "两边都没地址"

        notes = list(g.review_notes)
        if match_note:
            notes.append(match_note)
        if _LEFTOVER_JUNK_RE.search(g.name):
            notes.append("名称里仍有未清理的括号/书名号")
        # 地址来源相关说明(fallback 时附在「待复核」 列供人查看)
        if addr_note:
            notes.append(addr_note)
        # 教育文件候选信息(不论是否采用,都给人看)
        if best_entry is not None:
            tie_text = f"并列 {ties_count} 条" if ties_count >= 2 else "唯一"
            notes.append(f"教育文件候选:{best_entry.name}({tie_text},score={cand_score})")
        if notes:
            flagged += 1

        ws.append(
            [
                g.name,
                g.name_raw,
                g.committee_raw,
                g.type_raw,
                g.address_raw,
                json.dumps(g.alias, ensure_ascii=False) if g.alias else None,
                region_name,
                region_id,
                score,
                json.dumps(g.subarea, ensure_ascii=False) if g.subarea else None,
                1,
                json.dumps(chosen_addrs, ensure_ascii=False) if chosen_addrs else None,
                addr_source,
                cand_name,
                cand_addrs,
                cand_school,
                cand_score,
                "；".join(notes) if notes else None,
                ",".join(str(r) for r in g.source_rows),
            ]
        )

    widths = [24, 24, 20, 8, 36, 20, 20, 14, 10, 30, 6, 36, 14, 24, 36, 30, 8, 40, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # 备选 sheet:文件里所有 (小区, 地址, 学校) 平铺
    ws2 = wb.create_sheet("address_src")
    ws2.append(["小区名(原文)", "地址", "来源学校"])
    for e in addr_entries:
        # 每个地址一行,方便筛选/排序
        for a in e.addresses:
            ws2.append([e.name, a, e.school])
    for i, w in enumerate([28, 36, 40], start=1):
        ws2.column_dimensions[chr(64 + i)].width = w

    wb.save(OUT_XLSX)
    print(f"共 {len(groups)} 条小区记录 -> {OUT_XLSX}")
    print(f"居村委匹配: {matched} 成功 / {unmatched} 未匹配")
    print(f"待人工复核: {flagged} 条(见「待复核」列)")
    print(f"招生范围文件命中: {len(addr_entries)} 条原始记录(平铺到 address_src sheet)")


def main() -> None:
    raw_rows = load_source_rows()
    print(f"源数据 {len(raw_rows)} 行(含村委)")

    rows = expand_rows(raw_rows)
    print(f"筛选「类型=居委」并展开名称后: {len(rows)} 行")

    groups = group_rows(rows)
    print(f"按(名称, 所属居村委)分组后: {len(groups)} 条小区记录")

    committees = load_region_committees()
    print(f"region.json 居委会/村委会参考数据: {len(committees)} 条")
    matcher = RegionMatcher(committees)

    # 解析招生范围文件,作为地址备选源
    addr_entries = parse_address_file()
    print(f"招生范围文件解析: {len(addr_entries)} 条 (小区, 地址) 记录")

    build_output(groups, matcher, addr_entries)


if __name__ == "__main__":
    main()
